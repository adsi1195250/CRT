import json
from datetime import datetime, date

from django.core import serializers
from django.core.paginator import Paginator
from django.forms import model_to_dict
from django.http import HttpResponse, JsonResponse, response
from django.shortcuts import render, redirect, get_object_or_404

# Create your views here.
from django.urls import reverse_lazy
from django.views.generic import DetailView, DeleteView
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, Table, TableStyle

from CRT import settings
from main.forms import Historial_IOForms, Historial_IOForms_prueba, Historial_IO
from main.models import Trabajadores
from permiso_ausentismo.forms import PermisoAusentismoForms
from permiso_ausentismo.models import PermisoAusentismo
from excel_response import ExcelResponse


def ListarPermisoAusentismo(request):
    listar_permisos = PermisoAusentismo.objects.all()
    trabajador = ''
    cedula = request.GET.get('cedula')
    mes = request.GET.get('mes')
    if Trabajadores.objects.all().exists():
        hay_trabajador = 1
    else:
        hay_trabajador = 0
    # print(cedula)
    # print(mes)

    if cedula:
        listar_permisos = PermisoAusentismo.objects.filter(idTrabajador__cedula=cedula).order_by('-idPermisoAusentismo')
        trabajador = Trabajadores.objects.filter(cedula=cedula)
    else:
        listar_permisos = PermisoAusentismo.objects.filter().order_by('-idPermisoAusentismo')


    if mes and mes != '--':
        listar_permisos = PermisoAusentismo.objects.filter(periodoIncapacidadInicial__month=mes).order_by('periodoIncapacidadInicial')


    if mes and mes != '--' and cedula:
        listar_permisos = PermisoAusentismo.objects.filter(periodoIncapacidadInicial__month=mes, idTrabajador__cedula=cedula).order_by('periodoIncapacidadInicial')



    #print(listar_permisos.values())
    #print(listar_permisos[1].get_mes_evento_display())
    paginator = Paginator(listar_permisos, 10)

    page = request.GET.get('page')
    listar_permisos = paginator.get_page(page)
    permiso=['']
    if request.POST.get('excel'):
        mes = 'Mes'
        periodo_inicial = 'Periodo inicial'
        periodo_final = 'Periodo final'
        total_dias_incapacidad = 'Total días'
        total_horas = 'Tiempo total'
        codigo_diagnostico = 'Código de diagnostico'
        tipo_evento = 'Tipo de evento'
        observaciones = 'Observaciones'
        data = []
        data.append(
            [mes, periodo_inicial, periodo_final, total_dias_incapacidad, total_horas, tipo_evento, codigo_diagnostico,
             observaciones])
        cont = 0
        nombre_mes="ENERO"
        for x in listar_permisos.object_list.values():
            #x['mes_evento'] = listar_permisos[cont].get_mes_evento_display()
            nombre_mes = str(x["periodoIncapacidadInicial"]),
            x['totalDiasIncapacidad'] = str(x['totalDiasIncapacidad'])
            if x['totalDiasIncapacidad'] == '0':
                x['totalDiasIncapacidad'] = 'No aplica'
            #print(x['totalDiasIncapacidad'])
            x['tipo_evento'] = listar_permisos[cont].get_tipo_evento_display()
            #print(x['observaciones'])
            if x['horaInicial'] != None and x['horaFinal'] != None:
                print(str(x['horaInicial'])+'   '+str(x['horaFinal']))
                FMT = '%H:%M:%S'
                tdelta =datetime.strptime(str(x['horaFinal']),FMT) - datetime.strptime(str(x['horaInicial']),FMT)

            else:
                tdelta = 'No aplica'

            if x['observaciones'] == None:
                x['observaciones'] = 'Sin registro'
            if x['codigoDiagnostico'] == None or x['codigoDiagnostico'] == '':
                x['codigoDiagnostico'] = 'No aplica'
            if x['tipo_evento'][0:3] == 'A.C':
                x['tipo_evento'] = 'A.C'
            if x['tipo_evento'][0:3] == 'O.E':
                x['tipo_evento'] = 'O.E'
            if x['tipo_evento'][0:3] == 'A.T':
                x['tipo_evento'] = 'A.T'
            if x['tipo_evento'][0:3] == 'E.L':
                x['tipo_evento'] = 'E.L'

            formato = '%B'
            permiso=[
                x["periodoIncapacidadInicial"].strftime(formato),
                x["periodoIncapacidadInicial"],
                x["periodoIncapacidadFinal"],
                x["totalDiasIncapacidad"],
                tdelta,
                x["tipo_evento"],
                x["codigoDiagnostico"],
                x["observaciones"].lower(),
            ]

            #nombre_mes = listar_permisos[cont].get_mes_evento_display()
            cont = cont + 1

            data.append(permiso)

        nombre=''

        nombre_documento='Reporte_de_ausentismo'
        if trabajador:
            nombre = str(trabajador.first().nombres)
            nombre_documento = 'Reporte_de_ausentismo_'+nombre.replace(' ', '_')
            if mes:
                nombre_documento = 'Reporte_de_ausentismo_' + nombre.replace(' ', '_') + '_del_mes_' + permiso[0]

        if mes:
            nombre_documento = 'Reporte_de_ausentismo_del_mes_' + permiso[0]

        return ExcelResponse(data, nombre_documento, nombre)

    if request.POST.get('imprimir'):
        if trabajador:
            reponse = HttpResponse(content_type='application/pdf')
            # print(request.GET)
            reponse['Content-Disposition'] = 'filename=reporte.pdf'
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)

            # CABECERA DEL ARCHIVO
            archivo_imagen = settings.STATIC_ROOT + '/img/index.jpg'
            c.drawImage(archivo_imagen, 20, 750, 120, 80, preserveAspectRatio=True)
            c.setFont("Helvetica", 20)
            # Dibujamos una cadena en la ubicación X,Y especificada
            c.drawString(150, 800, u"Siete Colinas S.A.S")
            c.line(150, 795, 345, 795)
            c.setFont("Helvetica", 10)
            c.drawString(150, 780, u"REPORTE PERMISO DE AUSENTISMO")
            c.setFont("Helvetica", 15)
            c.drawString(430, 810, u"Colaborador:")
            c.setFont("Helvetica", 14)
            c.drawString(380, 794, str(trabajador.first()))
            c.setFont("Helvetica", 15)
            c.drawString(447, 773, u"Cedula:")
            c.setFont("Helvetica", 14)
            c.drawString(435, 756, cedula)

            #CABECERA DE LA TABLA
            styles = getSampleStyleSheet()
            styleBH = styles['Normal']
            styleBH.aligment = TA_CENTER
            styleBH.fontSize = 10

            mes = Paragraph('''Mes''',styleBH)
            periodo_inicial = Paragraph('''Periodo inicial''', styleBH)
            periodo_final = Paragraph('''Periodo final''', styleBH)
            total_dias_incapacidad = Paragraph('''Total días''', styleBH)
            total_horas = Paragraph('''Tiempo total''', styleBH)
            codigo_diagnostico = Paragraph('''Código de diagnostico''',styleBH)
            tipo_evento = Paragraph('''Tipo de evento''', styleBH)
            observaciones = Paragraph('''Observaciones''',styleBH)
            data = []
            data.append([mes,periodo_inicial,periodo_final,total_dias_incapacidad,total_horas,tipo_evento,codigo_diagnostico,observaciones])

            #TABLA
            styleN = styles['BodyText']
            styleN.aligment= TA_CENTER
            styleN.fontSize = 7

            high = 618

            p = ParagraphStyle('parrafos',
                               alignment=TA_LEFT,
                               fontSize=9,
                               fontName="Helvetica")
            cont = 0
            nombre_mes = ''
            for x in listar_permisos.object_list.values():
                #nombre_mes = x["periodoIncapacidadInicial"],
                x['totalDiasIncapacidad'] = str(x['totalDiasIncapacidad'])
                if x['totalDiasIncapacidad'] == '0':
                    x['totalDiasIncapacidad'] = 'No aplica'
                #print(x['totalDiasIncapacidad'])
                x['tipo_evento'] = listar_permisos[cont].get_tipo_evento_display()
                #print(x['observaciones'])
                if x['horaInicial'] != None and x['horaFinal'] != None:
                    print(str(x['horaInicial'])+'   '+str(x['horaFinal']))
                    FMT = '%H:%M:%S'
                    tdelta =datetime.strptime(str(x['horaFinal']),FMT) - datetime.strptime(str(x['horaInicial']),FMT)

                else:
                    tdelta = 'No aplica'

                if x['observaciones'] == None:
                    x['observaciones'] = 'Sin registro'
                if x['codigoDiagnostico'] == None or x['codigoDiagnostico'] == '':
                    x['codigoDiagnostico'] = 'No aplica'
                if x['tipo_evento'][0:3] == 'A.C':
                    x['tipo_evento'] = 'A.C'
                if x['tipo_evento'][0:3] == 'O.E':
                    x['tipo_evento'] = 'O.E'
                if x['tipo_evento'][0:3] == 'A.T':
                    x['tipo_evento'] = 'A.T'
                if x['tipo_evento'][0:3] == 'E.L':
                    x['tipo_evento'] = 'E.L'

                formato= '%B'
                permiso=[
                    #datetime.strptime(str(x["periodoIncapacidadInicial"]),formato),
                    x["periodoIncapacidadInicial"].strftime(formato),
                    x["periodoIncapacidadInicial"],
                    x["periodoIncapacidadFinal"],
                    x["totalDiasIncapacidad"],
                    tdelta,
                    x["tipo_evento"],
                    x["codigoDiagnostico"],
                    Paragraph(x["observaciones"].lower(),p),
                ]
                high = high - len(x['observaciones'])

                cont = cont + 1

                data.append(permiso)



            #TABLA
            width, height =A4

            table = Table(data,colWidths=[2.1 * cm,2.7 * cm,2.7 * cm,1.9 * cm,1.8 * cm,1.8 * cm,2.9 * cm,4.8 * cm,4.0 * cm,])

            table.setStyle(TableStyle([#Estilos de la tabla
                ('INNERGRID',(0,0),(-1,-1),0.25,colors.black),
                ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
                #('GRID',(6,1),(6,-1),2,colors.orange),
                ('ALIGN', (3, 0), (5, -1), 'CENTER'),

            ]))

            table.wrapOn(c,width,height)


            table.drawOn(c,5,high)
            print('-----')
            print(high)
            c.showPage()
            c.save()
            pdf = buffer.getvalue()
            buffer.close()
            reponse.write(pdf)
            return reponse
        elif mes:
            reponse = HttpResponse(content_type='application/pdf')
            # print(request.GET)
            reponse['Content-Disposition'] = 'filename=reporte.pdf'
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)

            # CABECERA DEL ARCHIVO
            archivo_imagen = settings.STATIC_ROOT + '/img/index.jpg'
            c.drawImage(archivo_imagen, 20, 750, 120, 80, preserveAspectRatio=True)
            c.setFont("Helvetica", 20)
            # Dibujamos una cadena en la ubicación X,Y especificada
            c.drawString(150, 800, u"Siete Colinas S.A.S")
            c.line(150, 795, 345, 795)
            c.setFont("Helvetica", 10)
            c.drawString(150, 780, u"REPORTE PERMISO DE AUSENTISMO")
            c.setFont("Helvetica", 15)
            c.drawString(430, 810, u"Mes:")

            c.setFont("Helvetica", 15)


            # CABECERA DE LA TABLA
            styles = getSampleStyleSheet()
            styleBH = styles['Normal']
            styleBH.aligment = TA_CENTER
            styleBH.fontSize = 10

            colaborador = Paragraph('''Colaborador''', styleBH)
            periodo_inicial = Paragraph('''Periodo inicial''', styleBH)
            periodo_final = Paragraph('''Periodo final''', styleBH)
            total_dias_incapacidad = Paragraph('''Total días''', styleBH)
            total_horas = Paragraph('''Tiempo total''', styleBH)
            codigo_diagnostico = Paragraph('''Código de diagnostico''', styleBH)
            tipo_evento = Paragraph('''Tipo de evento''', styleBH)
            observaciones = Paragraph('''Observaciones''', styleBH)
            data = []
            data.append([colaborador,periodo_inicial, periodo_final, total_dias_incapacidad, total_horas, tipo_evento,
                         codigo_diagnostico, observaciones])

            # TABLA
            styleN = styles['BodyText']
            styleN.aligment = TA_CENTER
            styleN.fontSize = 7

            high = 618

            p = ParagraphStyle('parrafos',
                               alignment=TA_LEFT,
                               fontSize=9,
                               fontName="Helvetica")
            cont = 0
            nombre_mes = ''
            for x in listar_permisos.object_list.values():
                # nombre_mes = x["periodoIncapacidadInicial"],

                if x['totalDiasIncapacidad'] == '0':
                    x['totalDiasIncapacidad'] = 'No aplica'
                # print(x['totalDiasIncapacidad'])
                x['tipo_evento'] = listar_permisos[cont].get_tipo_evento_display()
                x["idTrabajador_id"] = Trabajadores.objects.get(id=str(listar_permisos[cont].idTrabajador_id))


                # print(x['observaciones'])
                if x['horaInicial'] != None and x['horaFinal'] != None:
                    #print(str(x['horaInicial']) + '   ' + str(x['horaFinal']))
                    FMT = '%H:%M:%S'
                    tdelta = datetime.strptime(str(x['horaFinal']), FMT) - datetime.strptime(str(x['horaInicial']), FMT)

                else:
                    tdelta = 'No aplica'

                if x['observaciones'] == None:
                    x['observaciones'] = 'Sin registro'
                if x['codigoDiagnostico'] == None or x['codigoDiagnostico'] == '':
                    x['codigoDiagnostico'] = 'No aplica'
                if x['tipo_evento'][0:3] == 'A.C':
                    x['tipo_evento'] = 'A.C'
                if x['tipo_evento'][0:3] == 'O.E':
                    x['tipo_evento'] = 'O.E'
                if x['tipo_evento'][0:3] == 'A.T':
                    x['tipo_evento'] = 'A.T'
                if x['tipo_evento'][0:3] == 'E.L':
                    x['tipo_evento'] = 'E.L'

                formato = '%B'
                permiso = [
                    # datetime.strptime(str(x["periodoIncapacidadInicial"]),formato),
                    x["idTrabajador_id"],
                    x["periodoIncapacidadInicial"],
                    x["periodoIncapacidadFinal"],
                    x["totalDiasIncapacidad"],
                    tdelta,
                    x["tipo_evento"],
                    x["codigoDiagnostico"],
                    Paragraph(x["observaciones"].lower(), p),
                ]
                high = high - len(x['observaciones'])
                c.setFont("Helvetica", 14)
                c.drawString(420, 780, permiso[1].strftime(formato))
                cont = cont + 1

                data.append(permiso)

            # TABLA
            width, height = A4

            table = Table(data,
                          colWidths=[7.3 * cm,2.7 * cm, 2.7 * cm, 1.9 * cm, 1.8 * cm, 1.8 * cm, 2.9 * cm, 4.8 * cm,
                                     4.0 * cm, ])

            table.setStyle(TableStyle([  # Estilos de la tabla
                ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
                # ('GRID',(6,1),(6,-1),2,colors.orange),
                ('ALIGN', (3, 0), (5, -1), 'CENTER'),

            ]))

            table.wrapOn(c, width, height)

            table.drawOn(c, 5, high)
            print('-----')
            print(high)
            c.setPageSize((750, 850))
            c.showPage()

            c.save()
            pdf = buffer.getvalue()
            buffer.close()
            reponse.write(pdf)
            return reponse
    """
    if request.method == 'POST':
        if request.is_ajax():
            #report(request)
            return render(request, 'Trabajadores/trabajadores.html', context)
    """

    context = {
        'listar_permisos': listar_permisos,
        'trabajador': trabajador,
        'cedula':cedula,
        'mes':mes,
        'hay_trabajador':hay_trabajador,
    }
    return render(request, 'permiso_ausentismo/permiso_ausentismo.html', context)


def CrearPermisoAusentismo(request):
    form = PermisoAusentismoForms(request.POST or None)
    # --------------- Prueba de carga de datos xlsx ------------------- #
    context = {
        'form': form,
    }
    if request.method == 'POST':
        if request.is_ajax():
            # Always use get on request.POST. Correct way of querying a QueryDict.
            import openpyxl
            codigo = request.POST.get('codigo')
            documento = openpyxl.load_workbook('CIE/CIE10%s.xlsx' % codigo)
            lista = documento.get_sheet_names()
            sheet = documento.get_sheet_by_name(lista[0])
            a = []
            codigos = dict()

            # d,e=codigo.split('-')
            # password = request.POST.get('password')
            for filas in sheet.rows:
                for columnas in filas:
                    a.append(columnas.value)
                codigos[a[0]] = a[1]
                a = []
            """
            for filas in sheet[d:e]:
                for columna in filas:
                    a.append(columna.value)
                codigos[a[0]] = a[1]
                # print(a[0],a[1])
                a = []
            """
            # print(codigos)
            codigos.pop('Código')
            # print(codigos)
            data = {"codigo": codigo, "codigos": codigos}
            # codigos.clear()
            # print(codigos)
            documento.close()

            # Returning same data back to browser.It is not possible with Normal submit
            return JsonResponse(data)
        if form.is_valid():
            print(request.POST)
            if form.cleaned_data['observaciones'] == None:
                #print('awdad')
                form.initial={'observaciones':'Sin observaciones.'}
            instance = form.save(commit=False)
            instance.save()
            return redirect('listado_permiso_ausentismo')
        else:
            errors = form.errors
            # print(errors)
    return render(request, 'permiso_ausentismo/form_permiso_ausentismo.html', context)


"""
def DetallePermisoAusentismo(request,pk):
    permiso_ausentismo = PermisoAusentismo.objects.filter(idPermisoAusentismo=pk)
    context = {
        'object':permiso_ausentismo,
    }
    return render(request,'permiso_ausentismo/detalle_permiso_ausentismo.html',context)
"""


class DetallePermisoAusentismo(DetailView):
    model = PermisoAusentismo
    def get_queryset(self):
        queryset = super(DetallePermisoAusentismo, self).get_queryset()

        return queryset

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        FMT = '%H:%M:%S'
        totalHoras = ''
        #queryset = self.get_queryset().objects.filter(pk=self.get('object'))
       # print(queryset)
        #print(self.get_queryset().filter(pk=str(self.get_object())).values())
        for x in self.get_queryset().filter(pk=str(self.get_object())).values():
            if x['totalDiasIncapacidad'] == 0:
                if x['horaInicial'] != None and x['horaFinal'] != None:
                    totalHoras = str(datetime.strptime(str(x['horaFinal']), FMT) - datetime.strptime(
                        str(x['horaInicial']), FMT))
                else:
                    totalHoras = 'Registro incompleto'
        data['totalHoras'] = totalHoras
        return data
    template_name = 'permiso_ausentismo/detalle_permiso_ausentismo.html'


def prueba_ajax(request):
    if request.method == 'POST':
        # POST goes here . is_ajax is must to capture ajax requests. Beginner's pit.
        if request.is_ajax():
            # Always use get on request.POST. Correct way of querying a QueryDict.
            import openpyxl
            codigo = request.POST.get('codigo')

            documento = openpyxl.load_workbook('CIE10.xlsx')
            lista = documento.get_sheet_names()
            sheet = documento.get_sheet_by_name(lista[0])
            a = []
            codigos = dict()

            d, e = codigo.split('-')
            # password = request.POST.get('password')


            for filas in sheet[d:e]:
                for columna in filas:
                    a.append(columna.value)
                codigos[a[0]] = a[1]
                # print(a[0],a[1])
                a = []

            # print(codigos)
            codigos.pop('Código')
            print(codigos)
            data = {"codigo": codigo, "codigos": codigos}
            codigos.clear()
            documento.close()
            # Returning same data back to browser.It is not possible with Normal submit
            return JsonResponse(data)
            # Get goes here
    return render(request, 'permiso_ausentismo/tabs.html')


class EliminarPermisoAusentismo(DeleteView):
    model = PermisoAusentismo
    template_name = 'permiso_ausentismo/eliminar_permiso_ausentismo.html'
    success_url = reverse_lazy('listado_permiso_ausentismo')


def EditarPermisoAusentismo(request, pk):
    id = get_object_or_404(PermisoAusentismo, pk=pk)
    if request.method == 'POST':
        form = PermisoAusentismoForms(request.POST, instance=id)
        if request.is_ajax():
            # Always use get on request.POST. Correct way of querying a QueryDict.
            import openpyxl
            codigo = request.POST.get('codigo')
            documento = openpyxl.load_workbook('CIE/CIE10%s.xlsx' % codigo)
            lista = documento.get_sheet_names()
            sheet = documento.get_sheet_by_name(lista[0])
            a = []
            codigos = dict()

            # d,e=codigo.split('-')
            # password = request.POST.get('password')
            for filas in sheet.rows:
                for columnas in filas:
                    a.append(columnas.value)
                codigos[a[0]] = a[1]
                a = []
            """
            for filas in sheet[d:e]:
                for columna in filas:
                    a.append(columna.value)
                codigos[a[0]] = a[1]
                # print(a[0],a[1])
                a = []
            """
            # print(codigos)
            codigos.pop('Código')
            # print(codigos)
            data = {"codigo": codigo, "codigos": codigos}
            # codigos.clear()
            # print(codigos)
            documento.close()

            # Returning same data back to browser.It is not possible with Normal submit
            return JsonResponse(data)

        if form.is_valid():
            print(form.cleaned_data['observaciones'])
            if form.cleaned_data['observaciones'] == None:

                form.initial={'observaciones':'Sin observaciones.'}
            instance = form.save(commit=False)
            instance.save()
            return redirect('listado_permiso_ausentismo')
        else:
            print(form.errors)
    else:
        form = PermisoAusentismoForms(request.POST or None, instance=id)

    return render(request, 'permiso_ausentismo/form_permiso_ausentismo.html', {'form':form})


def reg_jorn_ajax(request):
    RegJornadaForm = Historial_IOForms_prueba(request.POST or None)
    if request.method == 'POST':
        # POST goes here . is_ajax is must to capture ajax requests. Beginner's pit.
        data = {}
        if request.is_ajax():
            # Always use get on request.POST. Correct way of querying a QueryDict.

            if request.POST.get('codigo'):
                codigo = request.POST.get('codigo')
                #print(codigo)
                trabajador = Trabajadores.objects.filter(CodigoBarras=codigo).values()
                #data = serializers.serialize('json',trabajador)
                data = list(trabajador)
                #print(data)
                return JsonResponse(data, safe=False)
            elif request.POST.get('codigo') != '':
                #RegJornadaForm = Historial_IOForms_prueba(request.POST or None, initial={'accion_jornada_hora': 'HEN','accion_jornada':request.POST['accion_jornada'],'id_trabajadores':request.POST['id_trabajador']})
                #RegJornadaForm(accion_jornada_hora='HEN',accion_jornada=request.POST['accion_jornada'],id_trabajadores=request.POST['id_trabajador'])
                #print(request.POST)

                trabajador = Trabajadores.objects.filter(id=request.POST['id_trabajadores'])
                #print(trabajador)
                RegJornadaForm = Historial_IOForms_prueba(request.POST)
                fecha = date.today().day
                fechaMes = date.today().month
                fechaAño = date.today().year
                fecha_trabajadores = Historial_IO.objects.all()
                lis = []
                for g in fecha_trabajadores:
                    for n in trabajador:
                        if g.fecha.day == fecha and g.fecha.month == fechaMes and g.fecha.year == fechaAño:
                            if g.id_trabajadores.cedula == n.cedula:
                                # print(g.accion_jornada, n.cedula)
                                lis.append(g.accion_jornada)
                #print(RegJornadaForm)
                if RegJornadaForm.is_valid():
                    instance = RegJornadaForm.save(commit=False)
                    accion_jornada = instance.accion_jornada
                    errors=[]
                    if 'SA' in lis:
                        errors.append("Usted ya registro SALIDA, no podrá ingresar mas datos.")
                        instance = RegJornadaForm.save(commit=False)
                    else:
                        if accion_jornada == 'EN':
                            instance.accion_jornada_hora = 'HEN'
                        elif accion_jornada == 'SA':
                            instance.accion_jornada_hora = 'HSA'
                        elif accion_jornada == 'DYI':
                            instance.accion_jornada_hora = 'HDYI'
                        elif accion_jornada == 'DYF':
                            instance.accion_jornada_hora = 'HDYF'
                        elif accion_jornada == 'DCI':
                            instance.accion_jornada_hora = 'HDCI'
                        elif accion_jornada == 'DCF':
                            instance.accion_jornada_hora = 'HDCF'
                        elif accion_jornada == 'PAI':
                            instance.accion_jornada_hora = 'HPAI'
                        elif accion_jornada == 'PAF':
                            instance.accion_jornada_hora = 'HPAF'
                        elif accion_jornada == 'ALI':
                            instance.accion_jornada_hora = 'HALI'
                        elif accion_jornada == 'ALF':
                            instance.accion_jornada_hora = 'HALF'

                    print(instance)
                    mensaje_full = {'guardado':'Guardado con exito'}
                    instance.save()
                    return JsonResponse(mensaje_full, safe=False)
                else:
                    print(RegJornadaForm.errors)
                    return JsonResponse({'errors':RegJornadaForm.errors}, safe=False)
            # Returning same data back to browser.It is not possible with Normal submit
            return JsonResponse(data,safe=False)
            # Get goes here
    return render(request, 'registrarJornada/reg_jorn_ajax.html',{'form':RegJornadaForm})
